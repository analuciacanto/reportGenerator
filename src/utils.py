from openpyxl.utils import get_column_letter

# First Tab


def setCellWidth(ws, min_row, min_col, max_col):

    column_widths = []

    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):

        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        ws.column_dimensions[col_name].width = value
