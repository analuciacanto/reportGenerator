from ctypes.wintypes import CHAR
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from utils import setCellWidth


def createOtherTabs(wb, users):
    for row in range(len(users)):  # PARA CADA USUÁRIO
        user = users[row]
        if user["forms"] != []:
            for form in range(len(user["forms"])):  # PARA CADA FORM DO USUÁRIO
                ws = wb.create_sheet(
                    title="Form-id" + str(user["forms"][form]["id"]))

                addContent(ws, user, form)
                setCellWidth(ws, 1, 1, ws.max_column)


def addContent(ws, user, form):

    header = ["Dec ID",	"UID",	"Nome do usuário",	"E-mail",	"CPF",	 "Empresa", "Departamento",	"Grupos a que pertence",
              "Gestor do usuário",	"Perfil do usuário",	'Status do usuário', "Data de cadastro do usuário", "Etapa atual da análise",	"Data da avaliação do Gestor",	"Preenchido Por", 	"Data do Preenchimento",	"Conclusão da análise",	"Última atualização feita por",	"Data da última atualização",	"Número de anexos na análise",	"Anotações internas",	"Parecer"]

    author = user["forms"][form]["author"]

    infos = ["--", user["id"], user["first_name"], user["email"],  user["number_id"],	 user["company"]["name"], user["department"]["name"],	"--",
             "--",	"--",	user["active"], user["created"],

             user["forms"][form]
             ["status"]["title"],	      user["forms"][form]
             ["approved_date"],	author["first_name"],	     user["forms"][form]
             ["created"],	     user["forms"][form]
             ["approved_date"],	"--",	"--",	"--",	"--",	"--"]

    otherInfos = []

    for i in range(32):
        try:

            if (isinstance(user["forms"][form]
                           ['questions'][str(i)]['value'], str)):

                header.append(user["forms"][form]
                              ['questions'][str(i)]['title'])
                infos.append(user["forms"][form]
                             ['questions'][str(i)]['value'])

            elif isinstance(user["forms"][form]
                            ['questions'][str(i)]['value'], list):

                header.append(user["forms"][form]
                              ['questions'][str(i)]['title'])
                infos.append("-")

                for value in user["forms"][form]['questions'][str(i)]['value']:

                    for question in value:
                        header.append(question['title'])
                        infos.append(question['value'])

        except KeyError:
            print("Não possui questão de índice: " + str(i))

    headerStyles(ws, header)
    ws.append(header)

    ws.append(infos)

    addStyles(ws, header)


def mergeCells(ws, startRow, endRow, startColumn, endColumn, initialCell, value, backgroundColor, fontColor, bold, fontSize):
    ws.merge_cells(start_row=startRow, start_column=startColumn,
                   end_row=endRow, end_column=endColumn)

    cell = ws[initialCell]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=backgroundColor)
    cell.alignment = Alignment(
        horizontal="center", vertical="center")
    cell.font = Font(bold=bold, color=fontColor,
                     size=fontSize, name='Calibri',)


def addHeaderColor(ws, row, firstColumn, lastColumn, color):
    for column in range(firstColumn, lastColumn):

        ws.cell(row, column).fill = PatternFill(start_color=color,
                                                end_color=color,
                                                fill_type='solid')
        ws.cell(row, column).font = Font(bold=False, color="00FFFFFF",
                                         size="12", name='Calibri',)


def addStyles(ws, header):
    addHeaderColor(ws, 2, 1, 13, '002f5496')
    addHeaderColor(ws, 2, 13, 23, '001f3864')
    addHeaderColor(ws, 2, 23, len(header) + 1, '004472c4')


def headerStyles(ws, header):
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    mergeCells(ws, 1, 1, 1, 12, "A1",  "Dados do usuário",
               "002f5496", "00FFFFFF", True, "15")

    mergeCells(ws, 1, 1, 13, 22,
               "M1", "Dados da análise",
               "001f3864", "00FFFFFF", True, "15")

    mergeCells(ws, 1, 1, 23,  len(header), "W1",  "Formulário",
               "004472c4", "00FFFFFF", True, "15")
