from math import ceil
from sqlite3 import Row
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from utils import setCellWidth


# First Tab
import re


def createSecondTab(wb, users):
    ws = wb.create_sheet(
        title="Form-id")
    addHeader(ws, users)
    setCellWidth(ws, 1, 1, ws.max_column)


def addHeader(ws, users):
    header = ["Dec ID",	"UID",	"Nome do usuário",	"E-mail",	"CPF",	 "Empresa", "Departamento",	"Grupos a que pertence",
              "Gestor do usuário",	"Perfil do usuário",	'Status do usuário', "Data de cadastro do usuário",
              "Etapa atual da análise",	"Data da avaliação do Gestor",	"Preenchido Por",
              "Data do Preenchimento",	"Conclusão da análise",	"Última atualização feita por",	"Data da última atualização",
                        "Número de anexos na análise",	"Anotações internas",	"Parecer",  "Você está preenchendo para outro colaborador? ",
              "Qual colaborador?", "Algum outro colaborador/Administrador da CCR estava presente?", "Adicionar Colaborador/Administrador", "Nome", "Cargo", "Divisão", "Unidade de Negócios",
              "Justifique:", "Dados dos Colaboradores/Administradores presentes na interação",   "Nome completo", "Cargo", "Divisão", "Unidade de Negócios:", "Dados dos Agentes Públicos presentes na interação",
              "Nome do Agente Público", "Órgão", "Área de atuação do Agente Público", "Data da interação:", "Local da interação:", "Horário de Início:",
                        "Horário de Término:", "Quem motivou a interação?", "Assunto discutido:", "Justifique:", "Qual?", "Qual o tema? ",
              "Houve formalização da interação em ata ou trocas de ofícios?", "Anexar documentos:",
              "A interação incluiu o custeio de algum tipo de hospitalidade (refeição, deslocamento, etc?",
              "Comentários?", "Anexar documentos (opcional): ",
              "Houve algum pedido, sinalização, indicação, requerimento ou conduta imprópria por qualquer dos presentes?", "Comentários:",
              "Anexe aqui (Opcional): ", "Houve algum pedido, sinalização, indicação ou requerimento para doação ou patrocínio de algum projeto ou evento?",
              "Comentários:", "Anexe aqui ( Opcional):", "Deseja informar algo não listado neste formulário?", "Comentários:", "Anexe aqui (Opcional):",
              "Declaro que as informações acima são verídicas completas e corretas, me responsabilizando pelo conteúdo nela contido", ]

    headerStyles(ws, header)
    ws.append(header)
    addContent(ws, users)
    addStyles(ws, header)


def getAnswer(index, questions):
    try:
        if (questions[str(index)]['value'] != [] or None):
            return questions[str(index)]['value']

        else:
            return "--"
    except KeyError:
        return "--"
    else:
        return "--"


def getAnswerList(index, questionIndex, questions, form, ws):
    try:
        question = questions[str(questionIndex)]['value']
        info = ""
        if (question != None):
            for i in range(len(question)):
                if i < len(question) - 1:
                    info = info + question[i][index]['value'] + "\n"
                else:
                    info = info + question[i][index]['value']
            return info

        else:
            return "--"

    except KeyError:
        return "--"
    else:
        return "--"


def addContent(ws, users):
    for user in range(len(users)):
        for form in range(len(users[user]["forms"])):
            questions = users[user]["forms"][form]['questions']
            author = users[user]["forms"][form]["author"]

            infos = ["--", users[user]["id"], users[user]["first_name"], users[user]["email"],  users[user]["number_id"],	 users[user]["company"]["name"], users[user]["department"]["name"],	"--",
                     "--",	"--",	users[user]["active"], users[user]["created"],

                     users[user]["forms"][form]
                     ["status"]["title"],	      users[user]["forms"][form]
                     ["approved_date"],	author["first_name"],	     users[user]["forms"][form]
                     ["created"],	     users[user]["forms"][form]
                     ["approved_date"],	"--",	"--",	"--",	"--",	"--", getAnswer(
                         0, questions), getAnswer(1, questions), getAnswer(2, questions), "--",

                     getAnswerList(0, 3, questions, user, ws),  getAnswerList(
                         1, 3, questions, user, ws),  getAnswerList(2, 3, questions, user, ws), getAnswerList(3, 3, questions, user, ws),

                     getAnswer(4, questions), "--", getAnswerList(0, 5, questions, user, ws), getAnswerList(
                         1, 5, questions, user, ws), getAnswerList(2, 5, questions, user, ws), getAnswerList(3, 5, questions, user, ws),
                     "--", getAnswerList(0, 6, questions, form, ws), getAnswerList(
                         1, 6, questions, user, ws), getAnswerList(2, 6, questions, user, ws),
                     getAnswer(7, questions), getAnswer(8, questions), getAnswer(
                         9, questions), getAnswer(10, questions),
                     getAnswer(11, questions),  getAnswer(
                         13, questions), getAnswer(14, questions), getAnswer(15, questions),
                     getAnswer(16, questions),
                     getAnswer(18, questions), getAnswer(19, questions), getAnswer(20, questions), getAnswer(21, questions), getAnswer(22, questions), getAnswer(23, questions), getAnswer(24, questions), getAnswer(25, questions), getAnswer(26, questions), getAnswer(27, questions), getAnswer(28, questions), getAnswer(29, questions), getAnswer(30, questions), getAnswer(31, questions), getAnswer(32, questions), ]

            ws.append(infos)
            ws.row_dimensions[user - 47].height = 50


def mergeCells(ws, startRow, endRow, startColumn, endColumn, initialCell, value, backgroundColor, fontColor, bold, fontSize):
    ws.merge_cells(start_row=startRow, start_column=startColumn,
                   end_row=endRow, end_column=endColumn)

    cell = ws[initialCell]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=backgroundColor)
    cell.alignment = Alignment(
        horizontal="center", vertical="center")
    cell.font = Font(bold=bold, color=fontColor,
                     size=fontSize, name='Calibri',)


def addHeaderColor(ws, row, firstColumn, lastColumn, color):
    for column in range(firstColumn, lastColumn):

        ws.cell(row, column).fill = PatternFill(start_color=color,
                                                end_color=color,
                                                fill_type='solid')
        ws.cell(row, column).font = Font(bold=False, color="00FFFFFF",
                                         size="12", name='Calibri',)


def addStyles(ws, header):
    addHeaderColor(ws, 2, 1, 13, '002f5496')
    addHeaderColor(ws, 2, 13, 23, '001f3864')
    addHeaderColor(ws, 2, 23, len(header) + 1, '004472c4')


def headerStyles(ws, header):
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    mergeCells(ws, 1, 1, 1, 12, "A1",  "Dados do usuário",
               "002f5496", "00FFFFFF", True, "15")

    mergeCells(ws, 1, 1, 13, 22,
               "M1", "Dados da análise",
               "001f3864", "00FFFFFF", True, "15")

    mergeCells(ws, 1, 1, 23,  len(header), "W1",  "Formulário",
               "004472c4", "00FFFFFF", True, "15")
