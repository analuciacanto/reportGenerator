from sqlite3 import Row
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from utils import setCellWidth

# First Tab
import re


def createFirstTab(wb, users):
    ws = wb.active
    ws.title = "Vigência"
    addHeader(ws)
    addContent(ws, users)
    setCellWidth(ws, 1, 1, ws.max_column)
    addHeaderStyles(ws)
    addContentStyles(ws, users)


def addHeader(ws):
    header = ["UID",	"Nome",	"E-mail",	"CPF",	"Empresa",	"Departamento",	"Grupos",
              "Perfil",	"Status",	"Dt Cadastro do usuário", "Formulário Preenchido?"]
    ws.append(header)


def addContent(ws, users):
    for row in range(len(users)):
        user = users[row]
        ws.append([user['id'], user['first_name'], user['email'], user['number_id'], user["company"]["name"], user["department"]
                  ["name"], "--", "--", "Ativo" if user['active'] else "Inativo", user['created'], "--" if user['forms'] == [] else "Preenchido"])


# Header


def addHeaderStyles(ws):

    for i in range(1, 12):

        ws.cell(1, i).fill = PatternFill(start_color='00003366',
                                         end_color='00003366',
                                         fill_type='solid')

        ws.cell(1, i).font = Font(bold=True, color="00FFFFFF",
                                  size="12", name='Calibri',)

        ws.cell(1, i).border = Border(left=Side(border_style='thin', color='00C0C0C0'),
                                      right=Side(border_style='thin',
                                                 color='00C0C0C0'),
                                      top=Side(border_style='thin',
                                               color='00C0C0C0'),
                                      bottom=Side(border_style='thin',
                                                  color='00C0C0C0'),)


def addContentStyles(ws, users):
    for row in range(2, len(users) + 2):
        for column in range(1, 12):
            if column == 11:
                if ws.cell(row, column).value == "Preenchido":
                    ws.cell(row, column).fill = PatternFill(start_color='00008000',
                                                            end_color='00008000',
                                                            fill_type='solid')
                else:
                    ws.cell(row, column).fill = PatternFill(start_color='00FF0000',
                                                            end_color='00FF0000',
                                                            fill_type='solid')
